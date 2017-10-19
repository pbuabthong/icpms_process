#! /usr/bin/env python3
# Copyright (c) 2017 Pai Pakpoom Buabthong

import openpyxl as opx
import numpy as np
from scipy import interpolate
import matplotlib.pyplot as plt
import math
import argparse
import os

def is_number(s):
    """Return True if the value is a number."""

    try:
        float(s)
        return True
    except ValueError:
        return False

def findFirstCell(sheet, search_str):
    """Return a coordinate tuple of the first containing the keyword.

    Usage: findFirstCell(sheet, 'Te')

    params:
    sheet: openpyxl.worksheet
        Excel worksheet object

    returns:
    coordinate(row, column)
        Excel coordinate of the cell
    """

    for row in sheet.iter_rows():
        for cell in row:
            if search_str in str(cell.value):
                return opx.utils.cell.coordinate_to_tuple(cell.coordinate)

def plotComparison(raw_list, std_list, lin_list, custom_fn):
    """Plot concentration vs cps for different interpolating method
    all lists are in increasing order.

    Usage: plotComparison(raw_list, std_list, lin_list, f)

    params:
    raw_list: List
        List of raw cps data
    std_list: List
        List of concentrations of standard samples (ug/L)
    lin_list: List
        List of concentrations of analyzed Cal-10 to Cal-1
    custom_fn: Scipy.interpolate.InterpolatedUnivariateSpline object
        or other function that can return y from custom_fn(x)
        Function of custom interpolated fitting
    """

    digits = round(math.log10(max(raw_list)))
    # generate a list with log interval from 1 to the max of raw_list
    xlog = np.logspace(1, digits)
    ylog = custom_fn(xlog)

    # plot comparing the fitting
    plt.loglog(raw_list, std_list, 'o', label='standard')
    plt.loglog(raw_list, lin_list, label='linear fitting from Agilent 8800')
    plt.loglog(xlog, ylog, label='interpolated univariate spline (k=1)')

    plt.xlabel('CPS')
    plt.ylabel('Concentration (ug/L)')

    plt.title("Different interpolations")

    plt.legend(loc = 'upper left')
    plt.show()

def main():
    """
    Add 1-degree interpolation smoothing spline for ICP-MS concentration.

    Modify the analyzed concentrations from Agilent ICP-MS 8800 QQQ.
    Use 1-degree smoothing spline for linear interpolation between each data points
    The fitting out of the standard range is a linear extrapolation
    from the minimum and maximum data points

    *this fixes one element at a time

    Example: python3 interp.icpms.py Te standard.xlsx raw_data.xlsx analyzed_data.xlsx

    Author: Pai Pakpoom Buabthong
    """

    parser = argparse.ArgumentParser(description = main.__doc__,
            formatter_class = argparse.RawDescriptionHelpFormatter)

    parser.add_argument('element', help = "Element to modify")
    parser.add_argument('std_filename', help = "Excel for standard concentration")
    parser.add_argument('raw_filename', help = "Excel for raw data in CPS")
    parser.add_argument('alz_filename', help = "Excel for analyzed data in ug/L")
    parser.add_argument('-v', "--verbosity", action = "store_true", default = False,
            help = "increase output verbosity")
    parser.add_argument('-p', help = "Print comparison graph", action = "store_true", default = False)
    parser.add_argument('-k', help = "Degree of smoothing spline", default = 1)
    args = parser.parse_args()

    global raw_filename, std_filename, alz_filename, v_value, p_value, k_value
    el = args.element
    raw_filename = args.raw_filename
    std_filename = args.std_filename
    alz_filename = args.alz_filename
    v_value = args.verbosity
    p_value = args.p
    k_value = args.k
    process(el, std_filename, raw_filename, alz_filename, v_value, p_value, k_value)


def process(el, std_filename, raw_filename, alz_filename, v_value,
        p_value, k_value):
    """
    Process data read from the given Excel files

    Usage: process('Te', 'std.xlsx', 'raw.xlsx', 'alz.xlsx', True, False, 1):

    params:
    el: str
        Name of the element to be modified
    std_filename: str
        Filename for standard concentration
    raw_filename: str
        Filename for raw data in CPS
    alz_filename: str
        Filename for analyzed data in ug/L
    v_value: bool
        Boolean value for increase verbosity
    p_value: bool
        Boolean value for graph plotting
    k_value: bool
        Degree of smoothing spline

    """
    print("Loading standard file:", std_filename)
    std = opx.load_workbook(std_filename, data_only = True)
    std_sheet = std.get_sheet_by_name(std.get_sheet_names()[0])
    keyword_coord = findFirstCell(std_sheet,'Final:')
    element_list_row = keyword_coord[0]-1

    std_list = []
    for element_col in std_sheet.iter_cols(
            min_row = element_list_row,
            min_col = (keyword_coord[1]+1)):
        if el in element_col[0].value:
            for cell in element_col[2:]:
                std_list.append(cell.value)
    std_list = std_list[::-1]

    print("Loading raw file:", std_filename)
    raw = opx.load_workbook(raw_filename, data_only = True)
    raw_sheet = raw.get_sheet_by_name(raw.get_sheet_names()[0])

    raw_list = []
    for raw_element_col in raw_sheet.iter_cols(max_row = 13):
        if el in str(raw_element_col[0].value):
            for cell in raw_element_col[3:]:
                raw_list.append(cell.value)

    f1d = interpolate.InterpolatedUnivariateSpline(raw_list, std_list, k=k_value)

    print("Loading analyzed file:", std_filename)
    alz = opx.load_workbook(alz_filename, data_only = True)
    alz_sheet = alz.get_sheet_by_name(alz.get_sheet_names()[0])

    samplename_column = findFirstCell(alz_sheet, 'Sample Name')[0]

    print("Converting...")
    name_list = []
    lin_list = []
    for raw_element_col in raw_sheet.iter_cols():
        if 'Sample Name' in str(raw_element_col[1].value):
            for cell in raw_element_col[2:]:
                name_list.append(cell.value)

        if (el in str(raw_element_col[0].value)):
            i = 0
            status = 'CPS'
            for cell in raw_element_col[2:]:
                if 0 < i <=10:
                    lin_list.append(alz_sheet[cell.coordinate].value)
                if is_number(cell.value):
                    try:
                        current_coord = cell.coordinate
                        value_new = float(f1d(cell.value))
                        value_old = alz_sheet[current_coord].value
                        if v_value :
                            print("Converting", name_list[i], status,
                                    ": (", cell.value, " )", value_old, "-->", value_new)
                        alz_sheet[current_coord].value = value_new
                    except:
                        print(name_list[i], ": cannot convert -- value: ", cell.value)
                else:
                    print(name_list[i], ": not float --  value: ", cell.value)
                i += 1
    save_filename = os.path.splitext(alz_filename)[0] + "_mod.xlsx"
    alz.save(save_filename)
    if p_value :
        plotComparison(raw_list, std_list, lin_list, f1d)

if __name__ == '__main__':
    main()
